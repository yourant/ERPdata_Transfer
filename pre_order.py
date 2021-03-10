import datetime
import pandas as pd
import pymysql
import xlrd
from openpyxl import Workbook
import retrieve_data
import os

day = (datetime.datetime.now() + datetime.timedelta(days=0)).strftime("%Y-%m-%d")
day_1 = (datetime.datetime.now() + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")
day_2 = (datetime.datetime.now() + datetime.timedelta(days=-2)).strftime("%Y-%m-%d")
day_3 = (datetime.datetime.now() + datetime.timedelta(days=-3)).strftime("%Y-%m-%d")
day_4 = (datetime.datetime.now() + datetime.timedelta(days=-4)).strftime("%Y-%m-%d")
day_5 = (datetime.datetime.now() + datetime.timedelta(days=-5)).strftime("%Y-%m-%d")
day_6 = (datetime.datetime.now() + datetime.timedelta(days=-6)).strftime("%Y-%m-%d")
day_7 = (datetime.datetime.now() + datetime.timedelta(days=-7)).strftime("%Y-%m-%d")
day_8 = (datetime.datetime.now() + datetime.timedelta(days=-8)).strftime("%Y-%m-%d")
day_9 = (datetime.datetime.now() + datetime.timedelta(days=-9)).strftime("%Y-%m-%d")
day_14 = (datetime.datetime.now() + datetime.timedelta(days=-14)).strftime("%Y-%m-%d")
day_add_1 = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%Y-%m-%d")

print(day)
conn_test = pymysql.connect(host='rm-2zeq92vooj5447mqzso.mysql.rds.aliyuncs.com',
                            port=3306, user='leiming',
                            passwd='vg4wHTnJlbWK8SY',
                            db="cider",
                            charset='utf8',
                            cursorclass=pymysql.cursors.DictCursor)
cur_test = conn_test.cursor()
conn = pymysql.connect(host='am-2zehm0z9s74088fy9131930.ads.aliyuncs.com',
                       port=3306, user='leiming',
                       passwd='MdLowUj5dEJLiu',
                       db="cider_fact",
                       charset='utf8',
                       cursorclass=pymysql.cursors.DictCursor)
cur = conn.cursor()


# end_day = '2021-02-19'


def get_stock_data(end_day):
    data = retrieve_data.retrieve_data(end_day)
    data_cp = data.get_cp_data()
    data_kc = data.get_kc_data()
    data_cgd = data.get_cgd_data()
    data_dd = data.get_dd_data()
    data_pd = data.get_pd_data()
    data_dzj = data.get_dzj_data()
    sheet1 = data_dd[['匹配SKU', '付款时间', '订单状态', '支付金额(CNY)', '数量', '缺货数量']]
    sheet2 = data_cgd[(data_cgd['状态'] == '采购中') & (data_cgd['仓库'] == '自建仓-坑头')]
    sheet2 = sheet2[['本地SKU', '供应商', '物品数量', '到货物品数量', '预计到货时间']]
    sheet3 = data_kc[data_kc['仓库'] == '坑头']
    sheet3 = sheet3[['本地sku', '图片URL', '合格总量', '合格锁定量', '均价（￥）']]
    sheet4 = data_dzj[['本地SKU', '当前入库单SKU到货数量', '当前入库单SKU已质检数量']]
    sheet5 = data_cp[data_cp['店铺'] == 'shopcider']
    sheet5 = sheet5[['Sku', 'Sku图片', '售卖状态', '库存策略', '发布时间']]
    sheet6 = data_pd[['本地SKU', '平台SKU']]
    sheet1 = sheet1.reset_index(drop=True)
    sheet2 = sheet2.reset_index(drop=True)
    sheet3 = sheet3.reset_index(drop=True)
    sheet4 = sheet4.reset_index(drop=True)
    sheet5 = sheet5.reset_index(drop=True)
    sheet6 = sheet6.reset_index(drop=True)
    writer = pd.ExcelWriter(os.getcwd() + '/{0}预期时间sop.xlsx'.format(end_day), engine='xlsxwriter')
    sheet1.to_excel(writer, sheet_name='order', index=False)
    sheet2.to_excel(writer, sheet_name='supply', index=False)
    sheet3.to_excel(writer, sheet_name='stock', index=False)
    sheet4.to_excel(writer, sheet_name='inspect', index=False)
    sheet5.to_excel(writer, sheet_name='online', index=False)
    sheet6.to_excel(writer, sheet_name='mapping', index=False)
    writer.save()


def write_data():
    data = xlrd.open_workbook(os.getcwd() + '/{0}预期时间sop.xlsx'.format(datetime.datetime.now().date()))
    for table_index in range(0, 5):
        print(table_index)
        sheet = data.sheets()[table_index]
        if table_index == 0:
            total_data_list = []
            db_table = 'banma_erp_order_new'
            title = 'sku,	pay_time, order_status,	total_fee, number, out_stock'
            table = data.sheets()[0]
            nrows = table.nrows
            for row in range(1, nrows):
                data_list = []
                i = 1
                for col in table.row(row):
                    if i == 2:
                        if col.value == '' or col.value is None:
                            data_list.append(None)
                        else:
                            data_list.append(xlrd.xldate.xldate_as_datetime(col.value, 0))
                    elif (i == 5 or i == 6) and col.value == '--':
                        data_list.append(0)
                    elif i == 5 or i == 6:
                        data_list.append(int(col.value))
                    else:
                        data_list.append(col.value)
                    i += 1
                total_data_list.append(tuple(data_list))
            cur.execute('truncate table banma_erp_order_new')
            sql = 'insert into {} ({}) '.format(db_table, title)
            cur.executemany(sql + '''values (%s,%s,%s,%s,%s,%s)''', total_data_list)
            conn.commit()
            # table_index += 1
        if table_index == 1:
            total_data_list = []
            db_table = 'banma_erp_supply_order_new'
            title = 'sku,	supplier_name,	total_stock,	arrive_stock, est_time'
            table = data.sheets()[table_index]
            nrows = table.nrows
            for row in range(1, nrows):
                data_list = []
                i = 1
                for col in table.row(row):
                    if i == 5:
                        if col.value == '' or col.value is None:
                            data_list.append(None)
                        else:
                            data_list.append(xlrd.xldate.xldate_as_datetime(col.value, 0))
                    elif (i == 3 or i == 4) and col.value == '--':
                        data_list.append(None)
                    else:
                        data_list.append(col.value)
                    i += 1
                # sql = '''insert into banma_erp_order (master_id,	order_id,	platform,	shop,	order_status,	platform_order_status,	shelve_status,	refund_status,	virtual_send,	buyer_comment,	is_comment,	comment,	checkout_time,	pay_time,	shipping_type,	pay_type,	trade_no,	buyer_mail,	buyer_id,	buyer_name,	country,	province,	city,	address,	address_2,	address_3,	mobile,	mail,	zip,	name,	amount_cny,	amount_usd,	currency,	refund_amount_cny,	refund_amount_usd,	product_title,	price,	number,	out_stock,	size,	erp_sku,	platform_sku,	ext_info)
                # values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                #                   %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'''
                # cur.execute(sql, data_list)
                # print(data_list)
                # conn.commit()
                # cur.close()
                # conn.close()
                total_data_list.append(tuple(data_list))
            print(total_data_list)
            cur.execute('truncate table banma_erp_supply_order_new')
            sql = 'insert into {} ({}) '.format(db_table, title)
            cur.executemany(sql + '''values (%s,%s,%s,%s,%s)''', total_data_list)
            conn.commit()
            print('success')
            print(123131)
            # table_index += 1
        if table_index == 2:
            total_data_list = []
            db_table = 'banma_erp_stock_new'
            title = 'sku,	product_url,	qualified_stock,	qualified_freeze_stock, avg_price'
            table = data.sheets()[table_index]
            nrows = table.nrows
            for row in range(1, nrows):
                data_list = []
                i = 1
                for col in table.row(row):
                    if (i == 3 or i == 4) and col.value == '--':
                        data_list.append(None)
                    elif i == 3 or i == 4:
                        data_list.append(int(col.value))
                    else:
                        data_list.append(col.value)
                    i += 1
                total_data_list.append(tuple(data_list))
            cur.execute('truncate table banma_erp_stock_new')
            sql = 'insert into {} ({}) '.format(db_table, title)
            cur.executemany(sql + '''values (%s,%s,%s,%s,%s)''', total_data_list)
            conn.commit()
        if table_index == 3:
            total_data_list = []
            db_table = 'banma_erp_inspection_new'
            title = 'sku,total_num,complete_num'
            table = data.sheets()[table_index]
            nrows = table.nrows
            for row in range(1, nrows):
                data_list = []
                i = 1
                for col in table.row(row):
                    if i == 2 or i == 3:
                        data_list.append(int(col.value))
                    else:
                        data_list.append(col.value)
                    i += 1
                total_data_list.append(tuple(data_list))
            cur.execute('truncate table banma_erp_inspection_new')
            sql = 'insert into {} ({}) '.format(db_table, title)
            cur.executemany(sql + '''values (%s,%s,%s)''', total_data_list)
            conn.commit()
            # table_index += 1
        if table_index == 4:
            total_data_list = []
            db_table = 'banma_erp_product_new'
            title = 'sku,product_url,is_alive,type,sale_time'
            table = data.sheets()[table_index]
            nrows = table.nrows
            print(nrows)
            for row in range(1, nrows):
                data_list = []
                i = 1
                for col in table.row(row):
                    if i == 5:
                        if col.value == '' or col.value is None:
                            data_list.append(None)
                        else:
                            data_list.append(xlrd.xldate.xldate_as_datetime(col.value, 0))
                    else:
                        data_list.append(col.value)
                    i += 1
                total_data_list.append(tuple(data_list))
            cur.execute('truncate table banma_erp_product_new')
            sql = 'insert into {} ({}) '.format(db_table, title)
            print(sql)
            cur.executemany(sql + '''values (%s,%s,%s,%s,%s)''', total_data_list)
            conn.commit()
        # 获取sku配对
        if table_index == 5:
            total_data_list = []
            db_table = 'banma_erp_match_new'
            title = 'sku,platform_sku'
            table = data.sheets()[table_index]
            nrows = table.nrows
            for row in range(1, nrows):
                data_list = []
                i = 1
                for col in table.row(row):
                    data_list.append(col.value)
                    i += 1
                total_data_list.append(tuple(data_list))
            cur.execute('truncate table banma_erp_match_new')
            sql = 'insert into {} ({}) '.format(db_table, title)
            cur.executemany(sql + '''values (%s,%s)''', total_data_list)
            conn.commit()
            # table_index += 1
    cur.execute('truncate table product_supply_daily_info')
    insert_sql = ''' insert into product_supply_daily_info (sku,est_time,sale_num, on_way_supply_num, out_stock, free_stock,able_stock_num,add_time,actual_able_stock,able_sale_days,day_1,day_2,day_3,day_4,day_5,day_6,day_7,inspection_num,product_url,day_7_avg,day_14_avg,sale_time)
select A.sku,est_time,sale_num, on_way_supply_num, out_stock, free_stock, if(free_stock is not null ,free_stock,0) + if(on_way_supply_num is not null and est_time is not null, on_way_supply_num, 0) - if(out_stock is not null,out_stock,0)  - (if(sale_num is not null, sale_num, 0) * if(diff_days is not null, diff_days, 0)) as able_stock_num,now(),
    if(free_stock is not null ,free_stock,0) + if(on_way_supply_num is not null and est_time is not null, on_way_supply_num, 0) - if(out_stock is not null,out_stock,0) as actual_able_stock,
    case when sale_num= 0 then null else cast((if(free_stock is not null, free_stock, 0)+ if(on_way_supply_num is not null
       and est_time is not null, on_way_supply_num, 0)  - if(out_stock is not null, out_stock, 0))  / if(sale_num is not null, sale_num, null) as bigint) end as able_sale_days,
     day_1,day_2,day_3,day_4,day_5,day_6,day_7,inspection_num,product_url,day_7_avg,day_14_avg,sale_time from 
    (select a.sku,b.est_time,b.on_way_supply_num,c.out_stock,c.sale_num,d.free_stock,DATEDIFF(b.est_time,now()) diff_days,day_1,day_2,day_3,day_4,day_5,day_6,day_7,g.inspection_num,case when g2.platform_sku is null then a.sku else g2.platform_sku end as platform_sku,c.day_7_avg,c.day_14_avg
     from 
    (select sku from banma_erp_order_new
    union 
    select sku from banma_erp_stock_new
    union
    select sku from banma_erp_supply_order_new) a

    left join 

    (select sku,case when date_add(date_format(est_time,'%Y-%m-%d'), interval 3 day) < now() then null else date_add(date_format(est_time,'%Y-%m-%d'), interval 4 day) end as est_time,max(supplier_name) as supplier_name,sum(if(total_stock is not null,total_stock,0) - if(arrive_stock is not null,arrive_stock,0)) as on_way_supply_num from banma_erp_supply_order_new
    group by sku,case when date_add(date_format(est_time,'%Y-%m-%d'), interval 3 day) < now() then null else date_add(date_format(est_time,'%Y-%m-%d'), interval 4 day) end) b on a.sku=b.sku

    left join 

    (select sku,cast(ceil(sum(case when pay_time>='{0}' and pay_time<'{1}' then number else 0 end) / 3) as bigint) as sale_num,sum(out_stock) as out_stock,
     sum(case when pay_time>='{2}' and pay_time<'{3}' then number else 0 end) as '{4}',
     sum(case when pay_time>='{5}' and pay_time<'{6}' then number else 0 end) as '{7}',
     sum(case when pay_time>='{8}' and pay_time<'{9}' then number else 0 end) as '{10}',
     sum(case when pay_time>='{11}' and pay_time<'{12}' then number else 0 end) as '{13}',
     sum(case when pay_time>='{14}' and pay_time<'{15}' then number else 0 end) as '{16}',
     sum(case when pay_time>='{17}' and pay_time<'{18}' then number else 0 end) as '{19}',
     sum(case when pay_time>='{20}' and pay_time<'{21}' then number else 0 end) as '{22}',
     cast(ceil(sum(case when pay_time>='{20}' and pay_time<'{3}' then number else 0 end) / 7) as bigint) as '{23}',
     cast(ceil(sum(case when pay_time>='{24}' and pay_time<'{3}' then number else 0 end) / 14) as bigint) as '{25}'from banma_erp_order_new
    where order_status <> '已取消' and total_fee>0
    group by sku) c on a.sku=c.sku

    left join 

    (select sku,max(product_url) as product_url,max(avg_price) as price,sum(if(qualified_stock is not null,qualified_stock,0)) - sum(if(qualified_freeze_stock is not null,qualified_freeze_stock,0)) as free_stock
    from banma_erp_stock_new
    group by sku) d on a.sku=d.sku

    left join

    (select sku,max(total_num) - max(complete_num) as inspection_num from banma_erp_inspection_new group by sku) g on a.sku = g.sku

    left join
    (select sku,max(platform_sku) as platform_sku from banma_erp_match_new group by sku) g2 on a.sku = g2.sku) A
join
(select sku,max(product_url) as product_url, max(sale_time) as sale_time from banma_erp_product_new where is_alive='上架' group by sku) g1 on g1.sku=A.platform_sku
     '''.format(day_3, day, day_1, day, 'day_1', day_2, day_1, 'day_2', day_3, day_2, 'day_3', day_4, day_3, 'day_4',
                day_5, day_4, 'day_5', day_6, day_5, 'day_6', day_7, day_6, 'day_7', 'day_7_avg', day_14, 'day_14_avg')
    print(insert_sql)
    cur.execute(insert_sql)
    conn.commit()


def get_data():
    sql = ''' select aaa.*,case when aaa.sku like 'gift%' then date_format('{0}','%Y-%m-%d %H:%i:%s') when aaa.now_able_stock > 0 then date_format('{0}','%Y-%m-%d %H:%i:%s') when b.est_time is null then date_format('2099-01-01','%Y-%m-%d %H:%i:%s') else b.est_time end as early_est_time,
    case when aaa.sale_num = 0 or aaa.sale_num is null then null else cast(aaa.able_stock / aaa.sale_num as bigint) end as able_sale_days from 
    (select * ,total_on_way_num + free_stock - out_stock + inspection_num as able_stock,
     free_stock - out_stock as now_able_stock
    from
    (select *,sum(on_way_supply_num) over(partition by sku order by est_time asc) as total_on_way_num 
    from 
    (select sku,case when est_time is null then date_format('2099-01-01','%Y-%m-%d %H:%i:%s') else est_time end as est_time,
     if(sale_num is null,0,sale_num) as sale_num,
     if(on_way_supply_num is null,0,on_way_supply_num) as on_way_supply_num,
     if(out_stock is null,0,out_stock) as out_stock,
     if(free_stock is null,0,free_stock) as free_stock,
     day_1,day_2,day_3,day_4,day_5,day_6,day_7,product_url,
     if(inspection_num is null,0,inspection_num) as inspection_num
     from product_supply_daily_info ) a) aa) aaa

    left join 

    (select sku,min(est_time) as est_time from 
    (select * ,total_on_way_num + free_stock - out_stock + inspection_num as able_stock
    from
    (select *,sum(on_way_supply_num) over(partition by sku order by est_time asc) as total_on_way_num 
    from 
    (select sku,case when est_time is null then date_format('2099-01-01','%Y-%m-%d %H:%i:%s') else est_time end as est_time,
     if(sale_num is null,0,sale_num) as sale_num,
     if(on_way_supply_num is null,0,on_way_supply_num) as on_way_supply_num,
     if(out_stock is null,0,out_stock) as out_stock,
     if(free_stock is null,0,free_stock) as free_stock,
     if(inspection_num is null,0,inspection_num) as inspection_num,
     day_1,day_2,day_3,day_4,day_5,day_6,day_7
     from product_supply_daily_info ) b) bb )bbb
    where able_stock > 0
    group by sku ) b
    on aaa.sku = b.sku'''.format(day_add_1)
    print(sql)
    cur.execute(sql)
    result = cur.fetchall()
    detail_list = []
    sku_list = []
    sku_time_list = []
    for r in result:
        r = list(r.values())
        detail_list.append(list(r))
        sku_only = [r[0], r[13], r[18], r[2], r[6], r[7], r[8], r[9], r[10], r[11], r[12]]
        if sku_only not in sku_list:
            sku_list.append(sku_only)
            sku_time_list.append([r[0], r[18]])
    # 写入商品分析表
    #     write_analysis_sql = '''insert into product_analysis
    #     select a.sku,a.sale_num,a.day_1,a.day_2,a.day_3,a.day_4,a.day_5,a.day_6,a.day_7,a.product_url,a.inspection_num,now_able_stock,max(total_on_way_num) as total_on_way_num,max(able_sale_days) as able_sale_days,day_7_avg,day_14_avg,sale_time from
    #  (select aaa.*,case when aaa.sku like 'gift%' then date_format('{0}','%Y-%m-%d %H:%i:%s') when aaa.now_able_stock > 0 then date_format('2021-01-12','%Y-%m-%d %H:%i:%s') when b.est_time is null then date_format('2099-01-01','%Y-%m-%d %H:%i:%s') else b.est_time end as early_est_time,
    #     case when aaa.sale_num = 0 or aaa.sale_num is null then null else cast(aaa.able_stock / aaa.sale_num as bigint) end as able_sale_days from
    #     (select * ,total_on_way_num + free_stock - out_stock + inspection_num as able_stock,
    #      free_stock - out_stock as now_able_stock
    #     from
    #     (select *,sum(on_way_supply_num) over(partition by sku order by est_time asc) as total_on_way_num
    #     from
    #     (select sku,case when est_time is null then date_format('2099-01-01','%Y-%m-%d %H:%i:%s') else est_time end as est_time,
    #      if(sale_num is null,0,sale_num) as sale_num,
    #      if(on_way_supply_num is null,0,on_way_supply_num) as on_way_supply_num,
    #      if(out_stock is null,0,out_stock) as out_stock,
    #      if(free_stock is null,0,free_stock) as free_stock,
    #      day_1,day_2,day_3,day_4,day_5,day_6,day_7,product_url,
    #      if(inspection_num is null,0,inspection_num) as inspection_num
    #      ,day_7_avg,day_14_avg,sale_time
    #      from product_supply_daily_info ) a) aa) aaa
    #
    #     left join
    #
    #     (select sku,min(est_time) as est_time from
    #     (select * ,total_on_way_num + free_stock - out_stock + inspection_num as able_stock
    #     from
    #     (select *,sum(on_way_supply_num) over(partition by sku order by est_time asc) as total_on_way_num
    #     from
    #     (select sku,case when est_time is null then date_format('2099-01-01','%Y-%m-%d %H:%i:%s') else est_time end as est_time,
    #      if(sale_num is null,0,sale_num) as sale_num,
    #      if(on_way_supply_num is null,0,on_way_supply_num) as on_way_supply_num,
    #      if(out_stock is null,0,out_stock) as out_stock,
    #      if(free_stock is null,0,free_stock) as free_stock,
    #      if(inspection_num is null,0,inspection_num) as inspection_num,
    #      day_1,day_2,day_3,day_4,day_5,day_6,day_7,day_7_avg,day_14_avg
    #      from product_supply_daily_info ) b) bb )bbb
    #     where able_stock > 0
    #     group by sku ) b
    #     on aaa.sku = b.sku) a
    # group by a.sku,a.sale_num,a.day_1,a.day_2,a.day_3,a.day_4,a.day_5,a.day_6,a.day_7,a.product_url,a.inspection_num,now_able_stock,day_7_avg,day_14_avg,sale_time '''.format(
    #         day_add_1)
    #     cur.execute('truncate table product_analysis')
    #     cur.execute(write_analysis_sql)
    #     conn.commit()
    # SKU最早发货时间写入测试库
    cur_test.execute('truncate table banma_erp_sku_time')
    sql_sku = 'insert into banma_erp_sku_time (sku, min_est_time) '
    cur_test.executemany(sql_sku + '''values (%s,%s)''', sku_time_list)
    conn_test.commit()
    # sql_sku = 'insert into banma_erp_sku_time_total (sku, min_est_time) '
    # cur_test.executemany(sql_sku + '''values (%s,%s)''', sku_time_list)
    # conn_test.commit()
    # 刷测试的SKU最早发货时间
    cur_test.execute('truncate table cc_cider.sku_time')
    sql_sku = 'insert into cc_cider.sku_time (sku, est_time) '
    cur_test.executemany(sql_sku + '''values (%s,%s)''', sku_time_list)
    conn_test.commit()
    sql = ''' update cc_cider.sku_time set est_time = null
            where est_time = '2099-01-01 00:00:00' '''
    cur_test.execute(sql)
    conn_test.commit()
    get_order_time_sql = ''' select B.* from 
(select a.order_no,max(min_est_time) as min_est_time from 
(select a.order_no,order_status,a.sku,min(case when b.min_est_time is null then '2099-01-01 00:00:00' else b.min_est_time end) as min_est_time,
case when b.min_est_time is null then '无法在在线商品列表中找到信息' else '在架' end as is_alive 
from erp_out_stock a
left join (select * from banma_erp_sku_time_total where add_time>='{0}' and add_time<'{1}') b 
on a.sku = b.sku
where a.add_time>=CURRENT_DATE()
and a.order_status <> '已取消'
group by a.order_no,order_status,a.sku) A
group by order_no
having max(min_est_time)='2099-01-01 00:00:00') A
join
(select a.order_no,order_status,a.sku,min(case when b.min_est_time is null then '2099-01-01 00:00:00' else b.min_est_time end) as min_est_time,
case when b.min_est_time is null then '不在架' else '在架' end as is_alive 
from erp_out_stock a
left join (select * from banma_erp_sku_time_total where add_time>='{0}' and add_time<'{1}') b 
on a.sku = b.sku
where a.add_time>=CURRENT_DATE()
and a.order_status <> '已取消'
group by a.order_no,order_status,a.sku) B on A.order_no=B.order_no
'''.format(day_2, day_1)
    get_order_time_num = ''' select est_time,count(distinct order_no) as order_num from 
                (select order_no,max(min_est_time) as est_time from 
                (select a.order_no,a.sku,min(case when b.min_est_time is null then '2099-01-01 00:00:00' else b.min_est_time end) as min_est_time 
                from erp_out_stock a
                left join (select * from banma_erp_sku_time_total where add_time>='{0}' and add_time<'{1}') b 
                on a.sku = b.sku
                where a.add_time>=CURRENT_DATE()
                and a.order_status <> '已取消'
                group by a.order_no,a.sku) a
                group by order_no) a 
                group by est_time '''.format(day_2, day_1)
    cur_test.execute(get_order_time_sql)
    order_details = cur_test.fetchall()
    order_detail_list = []
    for order_detail in order_details:
        order_detail = list(order_detail.values())
        order_detail_list.append([order_detail[0], order_detail[1], order_detail[2], order_detail[3], order_detail[4]])
    cur_test.execute(get_order_time_num)
    time_details = cur_test.fetchall()
    time_detail_list = []
    print(time_details)
    for time_detail in time_details:
        time_detail = list(time_detail.values())
        time_detail_list.append([time_detail[0], time_detail[1]])
    workbook = Workbook()
    # 写入批次明细
    title = ['SKU', '批次到货时间(2099表示遥遥无期)', '近3天日均销量', '批次在途件数', '当前所欠用户件数', '闲置库存', '近1天销售件数'
        , '近2天销售件数', '近3天销售件数', '近4天销售件数', '近5天销售件数', '近6天销售件数', '近7天销售件数', 'SKU链接', '质检数量', '批次累计在途件数'
        , '批次到货后富裕库存', '当前富裕库存', 'SKU最早发货时间(2099表示遥遥无期)', '可售天数']
    sheet = workbook.create_sheet(str('SKU批次明细'), index=0)
    sheet.append(title)
    for item in detail_list:
        sheet.append(item)
    # 写入SKU最早发货时间
    sku_title = ['SKU', 'SKU链接', 'SKU最早发货时间(2099表示遥遥无期)', '近3天日均销量', '近1天销售件数', '近2天销售件数', '近3天销售件数',
                 '近4天销售件数', '近5天销售件数', '近6天销售件数', '近7天销售件数']
    sheet = workbook.create_sheet(str('SKU最早发货时间明细'), index=1)
    sheet.append(sku_title)
    for item in sku_list:
        sheet.append(item)
    # 写入订单最早发货时间
    order_detail_title = ['订单号', '订单是否发货', 'SKU', 'SKU最早发货时间(2099表示遥遥无期)', 'SKU在架状态']
    sheet = workbook.create_sheet(str('无发货时间订单详情'), index=2)
    sheet.append(order_detail_title)
    for item in order_detail_list:
        sheet.append(item)
    # 写入订单最早发货时间分布
    order_detail_title = ['订单最早发货时间(2099表示遥遥无期)', '订单数', '占比']
    sheet = workbook.create_sheet(str('订单最早发货时间分布'), index=3)
    sheet.append(order_detail_title)
    Sum = 0
    for item in time_detail_list:
        Sum += item[1]
    for item in time_detail_list:
        item = item + tuple([item[1] / Sum])
        sheet.append(item)
    workbook.save(os.getcwd() + '/缺货{0}.xlsx'.format(datetime.datetime.now().date()))
    cur_test.close()
    conn_test.close()
    cur.close()
    conn.close()


if __name__ == '__main__':
    try:
        get_stock_data(str(datetime.datetime.now().date()))
        write_data()
        get_data()
    except Exception as e:
        print(e)
        quit()

